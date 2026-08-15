import csv
import io
import json
import threading
import time
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from flask import Flask, jsonify, render_template, request, send_file

app = Flask(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

state = {
    "running": False,
    "stop_requested": False,
    "page": 0,
    "items": 0,
    "logs": [],
    "results": [],
    "columns": [],
    "error": None,
    "last_url": "",
}

lock = threading.Lock()


def log(message):
    with lock:
        state["logs"].append(message)
        state["logs"] = state["logs"][-100:]


def classify_http_error(status_code):
    if status_code == 403:
        return "HTTP 403 Forbidden: the website refused the scraper request."
    if status_code == 404:
        return "HTTP 404 Not Found: the requested page does not exist."
    if status_code == 429:
        return "HTTP 429 Too Many Requests: the website is rate-limiting requests."
    if 500 <= status_code <= 599:
        return f"HTTP {status_code}: the target website reported a server error."
    return f"HTTP {status_code}: the target website rejected the request."


def fetch_soup(url):
    try:
        response = requests.get(
            url,
            headers=HEADERS,
            timeout=(10, 25),
            allow_redirects=True,
        )
        if not response.ok:
            message = classify_http_error(response.status_code)
            log(f"{message} URL: {url}")
            return None, message

        # requests normally detects encoding from headers; apparent_encoding
        # is only used when the server gives no useful charset.
        if not response.encoding or response.encoding.lower() == "iso-8859-1":
            response.encoding = response.apparent_encoding or "utf-8"

        return BeautifulSoup(response.text, "html.parser"), None

    except requests.exceptions.Timeout:
        message = "Request timed out while fetching the page."
        log(f"{message} URL: {url}")
        return None, message
    except requests.exceptions.RequestException as exc:
        message = f"Network error while fetching page: {exc}"
        log(message)
        return None, message


def clean_text(value):
    return " ".join(str(value or "").split())


def extract_value(node, mode, attribute="", base_url=""):
    if node is None:
        return ""

    mode = (mode or "text").lower()

    if mode == "text":
        return clean_text(node.get_text(" ", strip=True))

    if mode == "html":
        return str(node)

    if mode in {"href", "link"}:
        value = node.get("href", "")
        return urljoin(base_url, value) if value else ""

    if mode == "src":
        value = node.get("src", "")
        return urljoin(base_url, value) if value else ""

    if mode == "attribute":
        return clean_text(node.get(attribute, ""))

    return clean_text(node.get_text(" ", strip=True))


def extract_field(item, field, base_url):
    selector = str(field.get("selector", "")).strip()
    mode = str(field.get("mode", "text")).strip().lower()
    attribute = str(field.get("attribute", "")).strip()

    if selector:
        node = item.select_one(selector)
    else:
        node = item

    return extract_value(node, mode, attribute, base_url)


def _selector_for_node(node):
    """Build a compact CSS selector that is useful for repeated HTML cards."""
    if not node or not getattr(node, "name", None):
        return ""
    classes = [c for c in node.get("class", []) if c and len(c) < 50 and not c.startswith(("js-", "data-"))]
    # Prefer a small number of stable classes.
    if classes:
        return node.name + "".join("." + c.replace(" ", ".") for c in classes[:2])
    return node.name


def auto_detect_item_selector(soup):
    """Guess a repeating item/card selector from normal server-rendered HTML."""
    # Strong common patterns first.
    preferred = [
        "article.product_pod", "article", "li.product", 
        ".product_pod", ".product-item", ".product-card", 
        ".card", ".item", ".recipe", ".recipe-card", 
        ".post", ".article", "main article", "ul li"
    ]
    for selector in preferred:
        try:
            nodes = soup.select(selector)
            if 2 <= len(nodes) <= 500:
                return selector
        except Exception:
            pass

    # Look for repeated tag/class combinations.
    counts = {}
    for node in soup.find_all(["article", "li", "div", "section"], limit=2000):
        if not node.get("class"):
            continue
        selector = _selector_for_node(node)
        if not selector or selector in {"div", "li", "section", "article"}:
            continue
        try:
            count = len(soup.select(selector))
        except Exception:
            continue
        if 3 <= count <= 500:
            # Prefer selectors whose nodes contain a link/image/title-like element.
            sample = soup.select(selector)[:3]
            score = count
            for item in sample:
                if item.select_one("h1,h2,h3,h4,h5,a,img"):
                    score += 5
            counts[selector] = score
    if counts:
        return max(counts, key=counts.get)
    return "body"


def auto_detect_next_selector(soup):
    """Find a normal HTML pagination link, if present."""
    candidates = [
        'a[rel="next"]',
        'link[rel="next"]',
        'li.next a',
        'a.next',
        '.next a',
        'a[aria-label*="next" i]',
        'a[title*="next" i]',
    ]
    for selector in candidates:
        try:
            node = soup.select_one(selector)
            if node and node.get("href"):
                return selector
        except Exception:
            pass

    for a in soup.find_all("a"):
        label = clean_text(a.get_text(" ", strip=True)).lower()
        aria = str(a.get("aria-label", "")).lower()
        title = str(a.get("title", "")).lower()
        if a.get("href") and (label in {"next", "next page", "›", "→", ">"} or "next page" in aria or aria == "next" or "next" in title):
            return 'a[href="' + a.get("href", "").replace('"', '\\"') + '"]'
    return ""


def auto_detect_fields(item):
    """Infer useful columns from a detected card without requiring selectors."""
    fields = []
    def add(name, selector, mode="text", attribute=""):
        if not selector or any(f["name"] == name for f in fields):
            return
        if item.select_one(selector):
            fields.append({"name": name, "selector": selector, "mode": mode, "attribute": attribute})

    add("Title", "h1")
    add("Title", "h2")
    add("Title", "h3")
    add("Title", "h4")
    add("Title", ".title")
    add("Title", ".name")
    add("Price", ".price_color")
    add("Price", ".price")
    add("Price", "[class*='price']")
    add("Availability", ".availability")
    add("Availability", ".stock")
    add("Rating", ".star-rating", "attribute", "class")
    add("URL", "a", "href")
    add("Image", "img", "src")

    if not fields:
        fields = [{"name": "Text", "selector": "", "mode": "text", "attribute": ""}]
    return fields


def extract_data(soup, item_selector, fields, base_url):
    try:
        items = soup.select(item_selector)
    except Exception as exc:
        return [], f"Invalid item CSS selector: {exc}"

    if not items:
        return [], "Item selector returned zero results."

    rows = []
    for item in items:
        row = []
        for field in fields:
            row.append(extract_field(item, field, base_url))
        rows.append(row)

    return rows, None


def get_next_page(soup, current_url, next_selector):
    try:
        next_btn = soup.select_one(next_selector)
    except Exception as exc:
        return None, f"Invalid next-page CSS selector: {exc}"

    if not next_btn:
        return None, None

    href = next_btn.get("href")
    if not href:
        return None, "Next-page element was found but has no href."

    return urljoin(current_url, href), None


def validate_config(data):
    start_url = str(data.get("start_url", "")).strip()
    item_selector = str(data.get("item_selector", "")).strip()
    next_selector = str(data.get("next_selector", "")).strip()

    if not start_url:
        return None, "URL is required."

    parsed = urlparse(start_url)
    if parsed.scheme not in ("http", "https") or not parsed.netloc:
        return None, "Enter a valid http:// or https:// URL."

    try:
        max_pages = max(1, min(int(data.get("max_pages", 100)), 1000))
        delay = max(0.5, min(float(data.get("delay", 1.5)), 60))
    except (TypeError, ValueError):
        return None, "Invalid page limit or delay."

    raw_fields = data.get("fields") or []
    fields = []

    if raw_fields:
        for index, field in enumerate(raw_fields, start=1):
            name = clean_text(field.get("name", "")) or f"Field {index}"
            selector = str(field.get("selector", "")).strip()
            mode = str(field.get("mode", "text")).strip().lower()
            attribute = str(field.get("attribute", "")).strip()

            if mode not in {"text", "href", "src", "attribute", "html"}:
                return None, f"Unsupported extraction mode for '{name}'."

            if mode == "attribute" and not attribute:
                return None, f"Attribute name is required for '{name}'."

            fields.append({
                "name": name,
                "selector": selector,
                "mode": mode,
                "attribute": attribute,
            })
    else:
        # No fields supplied: infer useful columns from the first detected item.
        fields = []

    return {
        "start_url": start_url,
        "item_selector": item_selector,
        "next_selector": next_selector,
        "max_pages": max_pages,
        "delay": delay,
        "fields": fields,
    }, None


def scraper_thread(config):
    try:
        with lock:
            state.update({
                "base_url": config["start_url"],
                "results": [],
                "columns": [field["name"] for field in config["fields"]],
                "logs": [],
                "page": 0,
                "items": 0,
                "error": None,
                "running": True,
                "stop_requested": False,
                "last_url": config["start_url"],
            })

        url = config["start_url"]
        page = 1

        while url and page <= config["max_pages"]:
            with lock:
                if state["stop_requested"]:
                    log("Stopped by user.")
                    break
                state["page"] = page
                state["last_url"] = url

            log(f"Scraping page {page}: {url}")

            soup, error = fetch_soup(url)
            if soup is None:
                with lock:
                    state["error"] = error
                break

            if not config["item_selector"]:
                config["item_selector"] = auto_detect_item_selector(soup)
                log(f"Auto-detected item selector: {config['item_selector']}")

            if not config["fields"]:
                try:
                    first_item = soup.select_one(config["item_selector"])
                except Exception:
                    first_item = None
                config["fields"] = auto_detect_fields(first_item) if first_item else [{"name": "Text", "selector": "", "mode": "text", "attribute": ""}]
                with lock:
                    state["columns"] = [field["name"] for field in config["fields"]]
                log("Auto-detected extraction fields: " + ", ".join(field["name"] for field in config["fields"]))

            page_data, extraction_error = extract_data(
                soup,
                config["item_selector"],
                config["fields"],
                url,
            )

            if extraction_error:
                with lock:
                    state["error"] = extraction_error
                log(extraction_error)
                break

            with lock:
                state["results"].extend(page_data)
                state["items"] = len(state["results"])

            log(f"Found {len(page_data)} items on page {page}. Total: {state['items']}")

            next_selector = config["next_selector"] or auto_detect_next_selector(soup)
            if next_selector and not config["next_selector"]:
                config["next_selector"] = next_selector
                log(f"Auto-detected next-page selector: {next_selector}")
            next_url, next_error = get_next_page(
                soup, url, next_selector
            ) if next_selector else (None, None)
            if next_error:
                log(next_error)

            if not next_url:
                log("No next page found. Scraping complete.")
                break

            url = next_url
            page += 1

            if page <= config["max_pages"]:
                time.sleep(config["delay"])

        log(f"Finished. Total items: {state['items']}")

    except Exception as exc:
        with lock:
            state["error"] = f"Fatal error: {exc}"
        log(f"Fatal error: {exc}")
    finally:
        with lock:
            state["running"] = False


def get_snapshot():
    with lock:
        return {
            "running": state["running"],
            "page": state["page"],
            "items": state["items"],
            "logs": list(state["logs"][-50:]),
            "error": state["error"],
            "columns": list(state["columns"]),
            "results": [list(row) for row in state["results"][:500]],
        }


@app.route("/")
def index():
    return render_template("index.html")


@app.post("/start")
def start():
    if not request.is_json:
        return jsonify({"error": "JSON required"}), 400

    config, error = validate_config(request.get_json())
    if error:
        return jsonify({"error": error}), 400

    with lock:
        if state["running"]:
            return jsonify({"error": "A scraper is already running."}), 409

    thread = threading.Thread(target=scraper_thread, args=(config,), daemon=True)
    thread.start()
    return jsonify({"ok": True})


@app.post("/stop")
def stop():
    with lock:
        if state["running"]:
            state["stop_requested"] = True
            return jsonify({"ok": True, "message": "Stop requested."})
    return jsonify({"ok": True, "message": "No scraper is running."})


@app.get("/status")
def status():
    return jsonify(get_snapshot())


def build_csv():
    with lock:
        columns = list(state["columns"]) or ["Title"]
        rows = [list(row) for row in state["results"]]

    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(columns)
    writer.writerows(rows)

    mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    mem.seek(0)
    return mem


def build_json():
    with lock:
        columns = list(state["columns"]) or ["Title"]
        rows = [list(row) for row in state["results"]]

    records = [dict(zip(columns, row)) for row in rows]
    payload = json.dumps(records, ensure_ascii=False, indent=2).encode("utf-8")

    mem = io.BytesIO(payload)
    mem.seek(0)
    return mem


def build_xlsx():
    # Imported here so the application can still show a useful message if
    # a deployment forgot to install openpyxl.
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "XLSX export requires openpyxl. Add it to requirements.txt."
        ) from exc

    with lock:
        columns = list(state["columns"]) or ["Title"]
        rows = [list(row) for row in state["results"]]

    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    ws.append(columns)
    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    # Reasonable widths without creating huge columns.
    for column_cells in ws.columns:
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(value), 60))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(
            10, min(length + 2, 62)
        )

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return mem


@app.get("/download.csv")
def download_csv():
    try:
        mem = build_csv()
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    return send_file(
        mem,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name="scraped-data.csv",
    )


@app.get("/download.json")
def download_json():
    try:
        mem = build_json()
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    return send_file(
        mem,
        mimetype="application/json",
        as_attachment=True,
        download_name="scraped-data.json",
    )


@app.get("/download.xlsx")
def download_xlsx():
    try:
        mem = build_xlsx()
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="scraped-data.xlsx",
    )


# Keep the original /download URL working.
@app.get("/download")
def download_legacy():
    return download_csv()


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port, threaded=True)
